from __future__ import annotations

import os
from contextlib import asynccontextmanager
from datetime import date, datetime, timezone
from io import BytesIO
from zoneinfo import ZoneInfo

from apscheduler.schedulers.asyncio import AsyncIOScheduler
from fastapi import Depends, FastAPI, File, Form, HTTPException, Query, Request, UploadFile, WebSocket, WebSocketDisconnect
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook, load_workbook
from sqlalchemy import and_, func, inspect, or_, select, text
from sqlalchemy.orm import Session
from starlette.middleware.sessions import SessionMiddleware

from .database import Base, SessionLocal, engine, get_db
from .models import (
    ClientHeartbeat,
    ClientUpdateConfig,
    NotificationConfig,
    NotificationEvent,
    RegistrySyncLog,
    UserAccount,
    UserRole,
    VacationRecord,
    VacationStatus,
)
from .schemas import (
    ClientHeartbeatMonitorResponse,
    ClientHeartbeatResponse,
    ClientHeartbeatUpsert,
    ClientUpdateConfigResponse,
    ClientUpdateConfigUpdate,
    ClientUpdateManifestResponse,
    FIELD_HELP,
    NotificationConfigResponse,
    NotificationEventResponse,
    UserCreate,
    UserResponse,
    VacationCreate,
    VacationResponse,
    VacationUpdate,
)
from .services.auth import authenticate_user, ensure_default_admin, hash_password, normalize_username, verify_password
from .services.notifications import REGISTRY_UPDATE_TEXT, WebSocketHub, build_registry_update_message
from .services.vacation_service import run_registry_maintenance

APP_TZ = ZoneInfo(os.getenv("APP_TIMEZONE", "Europe/Moscow"))
SESSION_SECRET_KEY = os.getenv("SESSION_SECRET_KEY", "change-this-in-env")
SESSION_COOKIE_NAME = os.getenv("SESSION_COOKIE_NAME", "vacation_session")
SESSION_HTTPS_ONLY = os.getenv("SESSION_HTTPS_ONLY", "0") == "1"
CLIENT_HEARTBEAT_TOKEN = os.getenv("CLIENT_HEARTBEAT_TOKEN", "").strip()
hub = WebSocketHub()
scheduler = AsyncIOScheduler(timezone=APP_TZ)


def _read_positive_int_env(name: str, default: int) -> int:
    raw = os.getenv(name, str(default))
    try:
        value = int(raw)
    except (TypeError, ValueError):
        return default
    return value if value > 0 else default


RETURN_CHECK_INTERVAL_MINUTES = _read_positive_int_env("RETURN_CHECK_INTERVAL_MINUTES", 10)
ACTIVE_STATUSES = (VacationStatus.IN_VACATION, VacationStatus.SICK_LEAVE)
ACTIVE_STATUS_DB_VALUES = (
    VacationStatus.IN_VACATION,
    VacationStatus.SICK_LEAVE,
    VacationStatus.IN_VACATION.value,
    VacationStatus.SICK_LEAVE.value,
    VacationStatus.IN_VACATION.name,
    VacationStatus.SICK_LEAVE.name,
)
ACTIVE_STATUS_TEXT_NORMALIZED = {
    str(VacationStatus.IN_VACATION.value).strip().lower(),
    str(VacationStatus.SICK_LEAVE.value).strip().lower(),
    str(VacationStatus.IN_VACATION.name).strip().lower(),
    str(VacationStatus.SICK_LEAVE.name).strip().lower(),
}
SPECIAL_ALERT_LABEL = "ОСОБЫЙ СОТРУДНИК"
MAX_SPECIAL_NAMES_IN_NOTIFICATION = 5
MIN_NOTIFICATION_REPEAT_MINUTES = 0
MAX_NOTIFICATION_REPEAT_MINUTES = 1440
MIN_CLIENT_UPDATE_CHECK_MINUTES = 5
MAX_CLIENT_UPDATE_CHECK_MINUTES = 1440
CLIENT_ONLINE_WINDOW_SECONDS = 180
DEFAULT_CLIENT_VERSION = (os.getenv("DEFAULT_CLIENT_VERSION", "1.2.0") or "1.2.0").strip() or "1.2.0"
DEFAULT_UPDATES_BASE_PATH = "/static/updates"
DEFAULT_WINDOWS_UPDATE_FILENAME = "vacation-notifier-setup.exe"
DEFAULT_LINUX_DEB_UPDATE_FILENAME = "vacation-registry-notifier_latest_amd64.deb"
DEFAULT_LINUX_RPM_UPDATE_FILENAME = "vacation-registry-notifier-latest.x86_64.rpm"


def _is_admin(user: UserAccount | None) -> bool:
    return bool(user and user.role == UserRole.ADMIN and user.is_active)


def _is_editor(user: UserAccount | None) -> bool:
    return bool(user and user.role == UserRole.EDITOR and user.is_active)


def _is_client(user: UserAccount | None) -> bool:
    return bool(user and user.role in {UserRole.CLIENT, UserRole.VIEWER} and user.is_active)


def _can_manage_registry(user: UserAccount | None) -> bool:
    return bool(user and user.role in {UserRole.ADMIN, UserRole.EDITOR} and user.is_active)


def _role_label(user: UserAccount | None) -> str:
    if not user:
        return "Неизвестно"
    if user.role == UserRole.ADMIN:
        return "Администратор"
    if user.role == UserRole.EDITOR:
        return "Редактор"
    return "Клиент"


def _normalize_repeat_interval_minutes(value: int | str | None, default: int = 0) -> int:
    try:
        parsed = int(value) if value is not None else int(default)
    except (TypeError, ValueError):
        parsed = int(default)
    return max(MIN_NOTIFICATION_REPEAT_MINUTES, min(MAX_NOTIFICATION_REPEAT_MINUTES, parsed))


def _ensure_notification_config(db: Session) -> NotificationConfig:
    config = db.get(NotificationConfig, 1)
    if config:
        normalized = _normalize_repeat_interval_minutes(config.repeat_interval_minutes, 0)
        if normalized != config.repeat_interval_minutes:
            config.repeat_interval_minutes = normalized
            db.commit()
            db.refresh(config)
        return config

    config = NotificationConfig(id=1, repeat_interval_minutes=0)
    db.add(config)
    db.commit()
    db.refresh(config)
    return config


def _get_notification_repeat_interval_minutes(db: Session) -> int:
    config = _ensure_notification_config(db)
    return _normalize_repeat_interval_minutes(config.repeat_interval_minutes, 0)


def _normalize_update_check_interval_minutes(value: int | str | None, default: int = 60) -> int:
    try:
        parsed = int(value) if value is not None else int(default)
    except (TypeError, ValueError):
        parsed = int(default)
    return max(MIN_CLIENT_UPDATE_CHECK_MINUTES, min(MAX_CLIENT_UPDATE_CHECK_MINUTES, parsed))


def _normalize_update_version(value: str | None, default: str = DEFAULT_CLIENT_VERSION) -> str:
    normalized = str(value or "").strip()
    return normalized or default


def _split_version_parts(value: str) -> tuple:
    token = str(value or "").strip().lower().lstrip("v")
    if not token:
        return (0,)
    parts: list[tuple[int, str]] = []
    for chunk in token.replace("-", ".").split("."):
        piece = chunk.strip()
        if not piece:
            continue
        digits = "".join(ch for ch in piece if ch.isdigit())
        letters = "".join(ch for ch in piece if ch.isalpha())
        if digits:
            parts.append((int(digits), letters))
        else:
            parts.append((0, letters))
    return tuple(parts) if parts else (0,)


def _version_is_newer(candidate: str | None, current: str | None) -> bool:
    return _split_version_parts(candidate or "") > _split_version_parts(current or "")


def _ensure_client_update_config(db: Session) -> ClientUpdateConfig:
    config = db.get(ClientUpdateConfig, 1)
    if config:
        normalized_version = _normalize_update_version(config.latest_version, DEFAULT_CLIENT_VERSION)
        normalized_check = _normalize_update_check_interval_minutes(config.check_interval_minutes, 60)
        dirty = False
        if normalized_version != config.latest_version:
            config.latest_version = normalized_version
            dirty = True
        if normalized_check != config.check_interval_minutes:
            config.check_interval_minutes = normalized_check
            dirty = True
        if dirty:
            db.commit()
            db.refresh(config)
        return config

    config = ClientUpdateConfig(
        id=1,
        enabled=True,
        latest_version=DEFAULT_CLIENT_VERSION,
        windows_installer_url="",
        linux_deb_url="",
        linux_rpm_url="",
        check_interval_minutes=60,
    )
    db.add(config)
    db.commit()
    db.refresh(config)
    return config


def _serialize_client_update_config(config: ClientUpdateConfig) -> ClientUpdateConfigResponse:
    return ClientUpdateConfigResponse(
        enabled=bool(config.enabled),
        latest_version=_normalize_update_version(config.latest_version, DEFAULT_CLIENT_VERSION),
        windows_installer_url=_normalize_text(config.windows_installer_url),
        linux_deb_url=_normalize_text(config.linux_deb_url),
        linux_rpm_url=_normalize_text(config.linux_rpm_url),
        check_interval_minutes=_normalize_update_check_interval_minutes(config.check_interval_minutes, 60),
    )


def _build_default_client_update_urls(request: Request) -> dict[str, str]:
    base = str(request.base_url).rstrip("/")
    updates_base = DEFAULT_UPDATES_BASE_PATH.strip()
    if not updates_base.startswith("/"):
        updates_base = f"/{updates_base}"
    updates_base = updates_base.rstrip("/")

    return {
        "windows": f"{base}{updates_base}/{DEFAULT_WINDOWS_UPDATE_FILENAME}",
        "linux-deb": f"{base}{updates_base}/{DEFAULT_LINUX_DEB_UPDATE_FILENAME}",
        "linux-rpm": f"{base}{updates_base}/{DEFAULT_LINUX_RPM_UPDATE_FILENAME}",
    }


def _resolve_client_update_url(
    config: ClientUpdateConfig,
    platform_name: str,
    *,
    fallback_windows_url: str | None = None,
    fallback_linux_deb_url: str | None = None,
    fallback_linux_rpm_url: str | None = None,
) -> str | None:
    windows_url = _normalize_text(config.windows_installer_url) or _normalize_text(fallback_windows_url)
    linux_deb_url = _normalize_text(config.linux_deb_url) or _normalize_text(fallback_linux_deb_url)
    linux_rpm_url = _normalize_text(config.linux_rpm_url) or _normalize_text(fallback_linux_rpm_url)

    normalized = _normalize_text(platform_name).lower()
    if normalized == "windows":
        return windows_url or None
    if normalized == "linux-deb":
        return linux_deb_url or None
    if normalized == "linux-rpm":
        return linux_rpm_url or None
    if normalized == "linux":
        return linux_rpm_url or linux_deb_url or None
    return None

def _ensure_vacation_schema_updates() -> None:
    inspector = inspect(engine)
    try:
        columns = {column["name"] for column in inspector.get_columns("vacation_records")}
    except Exception:
        return

    if "is_special_employee" in columns:
        return

    dialect = engine.dialect.name.lower()
    if dialect == "postgresql":
        alter_sql = (
            "ALTER TABLE vacation_records "
            "ADD COLUMN IF NOT EXISTS is_special_employee BOOLEAN NOT NULL DEFAULT FALSE"
        )
    elif dialect == "sqlite":
        alter_sql = (
            "ALTER TABLE vacation_records "
            "ADD COLUMN is_special_employee BOOLEAN NOT NULL DEFAULT 0"
        )
    else:
        alter_sql = (
            "ALTER TABLE vacation_records "
            "ADD COLUMN is_special_employee BOOLEAN NOT NULL DEFAULT 0"
        )

    with engine.begin() as conn:
        conn.execute(text(alter_sql))
        try:
            conn.execute(
                text(
                    "CREATE INDEX IF NOT EXISTS ix_vacation_records_is_special_employee "
                    "ON vacation_records (is_special_employee)"
                )
            )
        except Exception:
            pass


def _ensure_client_heartbeat_schema_updates() -> None:
    inspector = inspect(engine)
    try:
        columns = {column["name"] for column in inspector.get_columns("client_heartbeats")}
    except Exception:
        return

    if "mac_address" not in columns:
        dialect = engine.dialect.name.lower()
        alter_sql = (
            "ALTER TABLE client_heartbeats ADD COLUMN IF NOT EXISTS mac_address VARCHAR(32)"
            if dialect == "postgresql"
            else "ALTER TABLE client_heartbeats ADD COLUMN mac_address VARCHAR(32)"
        )
        with engine.begin() as conn:
            conn.execute(text(alter_sql))

    with engine.begin() as conn:
        try:
            conn.execute(
                text(
                    "CREATE INDEX IF NOT EXISTS ix_client_heartbeats_mac_address "
                    "ON client_heartbeats (mac_address)"
                )
            )
        except Exception:
            pass


def _current_user_from_session(request: Request, db: Session) -> UserAccount | None:
    raw_id = request.session.get("user_id")
    if not raw_id:
        return None

    try:
        user_id = int(raw_id)
    except (TypeError, ValueError):
        request.session.pop("user_id", None)
        return None

    user = db.get(UserAccount, user_id)
    if not user or not user.is_active:
        request.session.pop("user_id", None)
        return None

    return user


def _require_user(request: Request, db: Session) -> UserAccount:
    user = _current_user_from_session(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Требуется вход в систему")
    return user


def _require_admin(request: Request, db: Session) -> UserAccount:
    user = _require_user(request, db)
    if user.role != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Недостаточно прав")
    return user


def _require_registry_editor(request: Request, db: Session) -> UserAccount:
    user = _require_user(request, db)
    if not _can_manage_registry(user):
        raise HTTPException(status_code=403, detail="Недостаточно прав для изменения реестра")
    return user


def _require_client_heartbeat_token(request: Request) -> None:
    if not CLIENT_HEARTBEAT_TOKEN:
        return
    token = _normalize_text(request.headers.get("x-client-token")) or ""
    if token != CLIENT_HEARTBEAT_TOKEN:
        raise HTTPException(status_code=403, detail="Недопустимый токен клиента.")


def require_user(request: Request, db: Session = Depends(get_db)) -> UserAccount:
    return _require_user(request, db)


def require_admin(request: Request, db: Session = Depends(get_db)) -> UserAccount:
    return _require_admin(request, db)


def require_registry_editor(request: Request, db: Session = Depends(get_db)) -> UserAccount:
    return _require_registry_editor(request, db)


def _maybe_user(request: Request, db: Session = Depends(get_db)) -> UserAccount | None:
    return _current_user_from_session(request, db)


def _redirect_to_login() -> RedirectResponse:
    return RedirectResponse(url="/login", status_code=303)


def _template_context(request: Request, user: UserAccount) -> dict:
    return {
        "request": request,
        "current_user": user,
        "is_admin": _is_admin(user),
        "is_editor": _is_editor(user),
        "is_client": _is_client(user),
        "can_manage_registry": _can_manage_registry(user),
        "role_label": _role_label(user),
    }


def _render_admin_page(
    request: Request,
    db: Session,
    current_user: UserAccount,
    *,
    error_text: str = "",
    success_text: str = "",
    status_code: int = 200,
):
    users = db.scalars(select(UserAccount).order_by(UserAccount.created_at.asc(), UserAccount.id.asc())).all()
    notification_config = _ensure_notification_config(db)
    update_config = _ensure_client_update_config(db)
    monitor_rows = _build_client_monitor_rows(db, latest_version=_normalize_update_version(update_config.latest_version, DEFAULT_CLIENT_VERSION))
    context = _template_context(request, current_user)
    context.update(
        {
            "users": users,
            "error_text": error_text,
            "success_text": success_text,
            "notification_repeat_interval_minutes": _normalize_repeat_interval_minutes(
                notification_config.repeat_interval_minutes,
                0,
            ),
            "notification_repeat_interval_max": MAX_NOTIFICATION_REPEAT_MINUTES,
            "client_update_config": _serialize_client_update_config(update_config),
            "client_monitor_rows": monitor_rows,
        }
    )
    return templates.TemplateResponse("admin.html", context, status_code=status_code)


def _render_account_page(
    request: Request,
    current_user: UserAccount,
    *,
    error_text: str = "",
    success_text: str = "",
    status_code: int = 200,
):
    context = _template_context(request, current_user)
    context.update({"error_text": error_text, "success_text": success_text})
    return templates.TemplateResponse("account.html", context, status_code=status_code)


def _validate_deputies(payload: VacationCreate | VacationUpdate) -> None:
    positions = {position.strip() for position in payload.employee_positions}
    for deputy in payload.deputies:
        if deputy.vacation_position.strip() not in positions:
            raise HTTPException(
                status_code=422,
                detail=(
                    f"Для заместителя '{deputy.deputy_full_name}' указана должность "
                    f"'{deputy.vacation_position}', которой нет в поле 'Должность/Должности'."
                ),
            )


def _apply_status_by_dates(record: VacationRecord) -> None:
    today = datetime.now(APP_TZ).date()
    if record.end_date < today:
        record.status = VacationStatus.FINISHED
        record.deactivated_at = datetime.now(timezone.utc)
    elif record.status == VacationStatus.FINISHED:
        record.deactivated_at = record.deactivated_at or datetime.now(timezone.utc)


def _normalize_text(value: str | None) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def _normalize_name_key(value: str) -> str:
    return _normalize_text(value).lower()


def _find_duplicate_vacation_record(
    db: Session,
    employee_full_name: str,
    start_date: date,
    end_date: date,
    *,
    exclude_id: int | None = None,
) -> VacationRecord | None:
    normalized_name = _normalize_name_key(employee_full_name)
    stmt = select(VacationRecord).where(
        VacationRecord.start_date == start_date,
        VacationRecord.end_date == end_date,
    )
    for record in db.scalars(stmt):
        if exclude_id is not None and record.id == exclude_id:
            continue
        if _normalize_name_key(record.employee_full_name) == normalized_name:
            return record
    return None


def _vacation_identity_key(
    employee_full_name: str,
    start_date: date,
    end_date: date,
    *,
    service: str = "",
) -> tuple[str, str, str, str]:
    return (
        _normalize_name_key(employee_full_name),
        start_date.isoformat(),
        end_date.isoformat(),
        _normalize_text(service).lower(),
    )


def _is_record_active_today(record: VacationRecord, today: date) -> bool:
    status_raw = record.status
    if isinstance(status_raw, VacationStatus):
        status_text = status_raw.value
    else:
        status_text = str(status_raw or "")
    is_active_status = status_text.strip().lower() in ACTIVE_STATUS_TEXT_NORMALIZED
    return (
        is_active_status
        and record.start_date <= today <= record.end_date
    )


def _deputies_are_placeholders(deputies: list[dict]) -> bool:
    if not deputies:
        return True
    return all(
        _normalize_text(str(dep.get("deputy_full_name", ""))).lower() in {"", "не назначен"}
        and _normalize_text(str(dep.get("deputy_actual_position", ""))).lower() in {"", "не назначен"}
        for dep in deputies
    )


def _split_cell_lines(value: object) -> list[str]:
    if value is None:
        return []

    text = str(value).replace("\r", "\n")
    parts: list[str] = []
    for line in text.split("\n"):
        for piece in line.split(";"):
            cleaned = _normalize_text(piece)
            if cleaned:
                parts.append(cleaned)
    return parts


def _parse_excel_status(value: object) -> VacationStatus:
    raw = _normalize_text(str(value) if value is not None else "")
    if not raw:
        return VacationStatus.IN_VACATION

    lowered = raw.lower()
    if lowered in {"в отпуске", "in vacation", "active"}:
        return VacationStatus.IN_VACATION
    if lowered in {"больничный лист", "больничный", "sick leave", "hospital"}:
        return VacationStatus.SICK_LEAVE
    if lowered in {"завершен", "завершён", "finished", "closed"}:
        return VacationStatus.FINISHED

    raise ValueError(f"Неизвестный статус: {raw}")


def _parse_excel_date(value: object, *, row_num: int, column_name: str) -> date:
    if value is None or _normalize_text(str(value)) == "":
        raise ValueError(f"Столбец '{column_name}' пуст.")

    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    raw = _normalize_text(str(value))
    date_formats = ("%Y-%m-%d", "%d.%m.%Y", "%d-%m-%Y", "%d/%m/%Y")

    for fmt in date_formats:
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue

    raise ValueError(f"Неверный формат даты в столбце '{column_name}' (строка {row_num}): {raw}")


def _parse_excel_special_employee(value: object) -> bool:
    raw = _normalize_text(str(value) if value is not None else "")
    if not raw:
        return False

    lowered = raw.lower()
    if lowered in {"1", "да", "yes", "true", "y", "истина", "особый", "vip"}:
        return True
    if lowered in {"0", "нет", "no", "false", "n", "обычный"}:
        return False

    raise ValueError(
        "Не удалось распознать значение поля 'Особый сотрудник'. "
        "Используйте: Да/Нет, True/False или 1/0."
    )


def _build_deputies_from_columns(
    employee_positions: list[str],
    deputy_names: list[str],
    deputy_actual_positions: list[str],
) -> list[dict]:
    if not employee_positions:
        raise ValueError("Не указана должность сотрудника.")
    if not deputy_names:
        raise ValueError("Не указан ни один заместитель.")

    if len(deputy_actual_positions) == 1 and len(deputy_names) > 1:
        deputy_actual_positions = deputy_actual_positions * len(deputy_names)
    if len(deputy_actual_positions) != len(deputy_names):
        raise ValueError("Количество заместителей и фактических должностей заместителей не совпадает.")

    if len(employee_positions) == len(deputy_names):
        vacation_positions = employee_positions
    elif len(employee_positions) == 1:
        vacation_positions = [employee_positions[0]] * len(deputy_names)
    elif len(deputy_names) == 1:
        vacation_positions = employee_positions
        deputy_names = [deputy_names[0]] * len(employee_positions)
        deputy_actual_positions = [deputy_actual_positions[0]] * len(employee_positions)
    else:
        raise ValueError(
            "Нельзя сопоставить заместителей с должностями. "
            "Укажите одинаковое количество должностей и заместителей."
        )

    deputies: list[dict] = []
    for idx, vacation_position in enumerate(vacation_positions):
        deputies.append(
            {
                "vacation_position": vacation_position,
                "deputy_full_name": deputy_names[idx],
                "deputy_actual_position": deputy_actual_positions[idx],
            }
        )
    return deputies


def _normalize_header(value: object) -> str:
    normalized = _normalize_text(str(value) if value is not None else "").lower().replace("ё", "е")
    for token in ('"', "'", "(", ")", "/", "\\", ":", ",", ".", "-", "№"):
        normalized = normalized.replace(token, " ")
    return " ".join(normalized.split())


EXCEL_COLUMN_LABELS: dict[str, str] = {
    "employee_full_name": "ФИО сотрудника",
    "is_special_employee": "Особый сотрудник",
    "employee_positions": "Должность/Должности",
    "status": "Статус",
    "service": "Услуга/Вид отпуска",
    "deputy_full_names": "Заместитель/и (ФИО)",
    "deputy_actual_positions": "Фактическая должность сотрудника (замещающего)",
    "start_date": "Дата начала отпуска",
    "end_date": "Дата окончания отпуска",
    "memo": "Памятка",
}

EXCEL_COLUMN_ALIASES: dict[str, list[str]] = {
    "employee_full_name": ["фио сотрудника", "фио", "сотрудник", "employee full name"],
    "is_special_employee": ["особый сотрудник", "особый", "special employee", "vip"],
    "employee_positions": ["должность должности", "должность", "employee positions"],
    "status": ["статус", "status"],
    "service": ["услуга", "вид отпуска", "service"],
    "deputy_full_names": ["заместитель и фио", "заместители фио", "заместитель фио", "deputy full name"],
    "deputy_actual_positions": [
        "фактическая должность сотрудника замещающего",
        "фактическая должность заместителя",
        "deputy actual position",
    ],
    "start_date": ["дата начала отпуска", "дата начала", "start date"],
    "end_date": [
        "дата окончания отпуска",
        "дата окончания",
        "дата компенсации окончания отпуска",
        "end date",
    ],
    "memo": ["памятка", "memo", "comment", "комментарий"],
}

EXCEL_REQUIRED_COLUMNS = ("employee_full_name", "employee_positions", "start_date", "end_date")


def _resolve_excel_column_indexes(header_row: tuple[object, ...]) -> dict[str, int]:
    normalized_headers = [_normalize_header(value) for value in header_row]
    indexes: dict[str, int] = {}

    for key, options in EXCEL_COLUMN_ALIASES.items():
        normalized_options = {_normalize_header(option) for option in options}
        for idx, header in enumerate(normalized_headers):
            if header and header in normalized_options:
                indexes[key] = idx
                break
    return indexes


def _find_excel_header_row(all_rows: list[tuple[object, ...]], *, scan_limit: int = 50) -> tuple[int, dict[str, int]]:
    if not all_rows:
        raise ValueError("В файле отсутствуют строки.")

    required_keys = set(EXCEL_REQUIRED_COLUMNS)
    best_row_num = 1
    best_indexes: dict[str, int] = {}
    best_score = -1

    for row_num, row in enumerate(all_rows[:scan_limit], start=1):
        indexes = _resolve_excel_column_indexes(row)
        score = len(required_keys.intersection(indexes.keys()))
        if score > best_score:
            best_score = score
            best_row_num = row_num
            best_indexes = indexes
        if score == len(required_keys):
            return row_num, indexes

    missing_labels = [EXCEL_COLUMN_LABELS[key] for key in EXCEL_REQUIRED_COLUMNS if key not in best_indexes]
    raise ValueError(
        "Не удалось распознать заголовки Excel. "
        "Отсутствуют обязательные столбцы: "
        + ", ".join(missing_labels)
        + ". Проверьте файл или используйте шаблон, выгруженный из системы."
    )


def _build_special_registry_alert_message(special_names: list[str]) -> str:
    preview = special_names[:MAX_SPECIAL_NAMES_IN_NOTIFICATION]
    names_text = ", ".join(preview)
    suffix = ""
    hidden_count = len(special_names) - len(preview)
    if hidden_count > 0:
        suffix = f" Дополнительно сотрудников: {hidden_count}."

    if len(special_names) == 1:
        return f"{SPECIAL_ALERT_LABEL}: {names_text} ушел(ушла) в отпуск.{suffix}"
    return f"ОСОБЫЕ СОТРУДНИКИ: {names_text} ушли в отпуск.{suffix}"


def _collect_special_active_employee_names(db: Session, changed_ids: list[int]) -> list[str]:
    if not changed_ids:
        return []

    today = datetime.now(APP_TZ).date()
    rows = db.scalars(
        select(VacationRecord.employee_full_name)
        .where(
            VacationRecord.id.in_(changed_ids),
            VacationRecord.is_special_employee.is_(True),
            VacationRecord.status.in_(ACTIVE_STATUS_DB_VALUES),
            VacationRecord.start_date <= today,
            VacationRecord.end_date >= today,
        )
        .order_by(VacationRecord.employee_full_name.asc())
    ).all()

    seen: set[str] = set()
    ordered_names: list[str] = []
    for raw_name in rows:
        normalized_name = _normalize_name_key(raw_name)
        if not normalized_name or normalized_name in seen:
            continue
        seen.add(normalized_name)
        ordered_names.append(_normalize_text(raw_name))
    return ordered_names


def _build_registry_notification_message(db: Session, changed_ids: list[int]) -> tuple[str, list[str]]:
    special_names = _collect_special_active_employee_names(db, changed_ids)
    if not special_names:
        return REGISTRY_UPDATE_TEXT, []
    special_line = _build_special_registry_alert_message(special_names)
    return f"{REGISTRY_UPDATE_TEXT}\n{special_line}", special_names


def _create_registry_notification_event(
    db: Session,
    changed_ids: list[int],
    *,
    message: str = REGISTRY_UPDATE_TEXT,
) -> NotificationEvent:
    event = NotificationEvent(
        event_type="registry_updated",
        message=message,
        changed_ids=changed_ids,
    )
    db.add(event)
    db.commit()
    db.refresh(event)
    return event


def _create_registry_notification_for_changes(db: Session, changed_ids: list[int]) -> tuple[NotificationEvent, list[str]]:
    message, special_names = _build_registry_notification_message(db, changed_ids)
    event = _create_registry_notification_event(db, changed_ids, message=message)
    return event, special_names


def _add_sync_log(
    db: Session,
    *,
    action_type: str,
    payload: dict,
    source_name: str | None = None,
    actor_username: str | None = None,
) -> None:
    log = RegistrySyncLog(
        action_type=action_type,
        source_name=source_name,
        actor_username=actor_username,
        payload=payload,
    )
    db.add(log)


def _serialize_notification(event: NotificationEvent, repeat_interval_minutes: int) -> NotificationEventResponse:
    return NotificationEventResponse(
        id=event.id,
        type=event.event_type,
        message=event.message,
        changed_ids=event.changed_ids or [],
        created_at=event.created_at,
        repeat_interval_minutes=_normalize_repeat_interval_minutes(repeat_interval_minutes, 0),
    )


async def _broadcast_registry_update(
    changed_ids: list[int],
    event: NotificationEvent,
    *,
    special_employee_names: list[str] | None = None,
    repeat_interval_minutes: int | None = None,
) -> None:
    effective_repeat_interval = repeat_interval_minutes
    if effective_repeat_interval is None:
        with SessionLocal() as settings_db:
            effective_repeat_interval = _get_notification_repeat_interval_minutes(settings_db)

    await hub.broadcast(
        build_registry_update_message(
            changed_ids,
            notification_id=event.id,
            changed_at=event.created_at,
            message=event.message,
            special_employee_names=special_employee_names,
            repeat_interval_minutes=_normalize_repeat_interval_minutes(effective_repeat_interval, 0),
        )
    )


async def _run_cleanup_and_notify() -> None:
    changed_ids: list[int] = []
    activated_ids: list[int] = []
    should_notify = False
    event: NotificationEvent | None = None
    special_names: list[str] = []

    with SessionLocal() as db:
        maintenance = run_registry_maintenance(db, datetime.now(APP_TZ).date(), retention_days=365)
        changed_ids = maintenance.changed_ids
        activated_ids = sorted(set(maintenance.activated_ids))
        return_messages = _build_return_messages(db, maintenance.finalized_ids)
        # Notify users only when a record appears in current vacations table.
        should_notify = bool(activated_ids)
        if should_notify:
            _add_sync_log(
                db,
                action_type="scheduled_sync",
                payload={
                    "activated_count": len(maintenance.activated_ids),
                    "finalized_count": len(maintenance.finalized_ids),
                    "deleted_count": len(maintenance.deleted_ids),
                    "changed_ids": changed_ids,
                    "activated_ids": activated_ids,
                    "messages": return_messages,
                },
            )
            db.commit()
        if should_notify:
            event, special_names = _create_registry_notification_for_changes(db, activated_ids)

    if should_notify and event:
        await _broadcast_registry_update(activated_ids, event, special_employee_names=special_names)


async def _run_return_deadline_notify() -> None:
    changed_ids: list[int] = []
    activated_ids: list[int] = []
    should_notify = False
    event: NotificationEvent | None = None
    special_names: list[str] = []

    with SessionLocal() as db:
        maintenance = run_registry_maintenance(
            db,
            datetime.now(APP_TZ).date(),
            retention_days=365,
            include_activations=False,
        )
        changed_ids = maintenance.changed_ids
        activated_ids = sorted(set(maintenance.activated_ids))
        # For this checker we do not create user notifications: no new active rows appear here.
        should_notify = bool(activated_ids)
        if should_notify:
            return_messages = _build_return_messages(db, maintenance.finalized_ids)
            _add_sync_log(
                db,
                action_type="scheduled_return_sync",
                payload={
                    "finalized_count": len(maintenance.finalized_ids),
                    "deleted_count": len(maintenance.deleted_ids),
                    "changed_ids": changed_ids,
                    "activated_ids": activated_ids,
                    "messages": return_messages,
                },
            )
            db.commit()
            event, special_names = _create_registry_notification_for_changes(db, activated_ids)

    if should_notify and event:
        await _broadcast_registry_update(activated_ids, event, special_employee_names=special_names)


@asynccontextmanager
async def lifespan(_: FastAPI):
    Base.metadata.create_all(bind=engine)
    _ensure_vacation_schema_updates()
    _ensure_client_heartbeat_schema_updates()

    with SessionLocal() as db:
        ensure_default_admin(db)
        _ensure_notification_config(db)
        _ensure_client_update_config(db)
        _cleanup_client_heartbeat_duplicates(db)
    _ensure_client_heartbeat_unique_indexes()

    await _run_cleanup_and_notify()

    if not scheduler.running:
        scheduler.add_job(_run_cleanup_and_notify, "cron", hour=0, minute=0, second=0, id="midnight_cleanup")
        scheduler.add_job(
            _run_return_deadline_notify,
            "interval",
            minutes=RETURN_CHECK_INTERVAL_MINUTES,
            id="return_deadline_check",
        )
        scheduler.start()

    yield

    if scheduler.running:
        scheduler.shutdown(wait=False)


app = FastAPI(
    title="Реестр отпусков",
    description="Клиент-серверное приложение кадрового отдела",
    lifespan=lifespan,
)

app.add_middleware(
    SessionMiddleware,
    secret_key=SESSION_SECRET_KEY,
    session_cookie=SESSION_COOKIE_NAME,
    same_site="lax",
    https_only=SESSION_HTTPS_ONLY,
)

app.mount("/static", StaticFiles(directory="app/static"), name="static")
templates = Jinja2Templates(directory="app/templates")


@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request, user: UserAccount | None = Depends(_maybe_user)):
    if user:
        target = "/" if _can_manage_registry(user) else "/active"
        return RedirectResponse(url=target, status_code=303)

    return templates.TemplateResponse("login.html", {"request": request, "error_text": ""})


@app.post("/login", response_class=HTMLResponse)
def login_submit(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    db: Session = Depends(get_db),
):
    user = authenticate_user(db, username=username, password=password)
    if not user:
        return templates.TemplateResponse(
            "login.html",
            {"request": request, "error_text": "Неверный логин или пароль."},
            status_code=401,
        )

    request.session["user_id"] = user.id
    target = "/" if _can_manage_registry(user) else "/active"
    return RedirectResponse(url=target, status_code=303)


@app.post("/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse(url="/login", status_code=303)


@app.get("/account", response_class=HTMLResponse)
def account_page(
    request: Request,
    db: Session = Depends(get_db),
):
    user = _current_user_from_session(request, db)
    if not user:
        return _redirect_to_login()
    if not _is_admin(user):
        return RedirectResponse(url="/active", status_code=303)
    return _render_account_page(request, user)


@app.post("/account/password", response_class=HTMLResponse)
def change_own_password(
    request: Request,
    current_password: str = Form(...),
    new_password: str = Form(...),
    new_password_confirm: str = Form(...),
    db: Session = Depends(get_db),
):
    user = _current_user_from_session(request, db)
    if not user:
        return _redirect_to_login()
    if not _is_admin(user):
        return RedirectResponse(url="/active", status_code=303)

    if not verify_password(current_password, user.password_hash):
        return _render_account_page(
            request,
            user,
            error_text="Текущий пароль указан неверно.",
            status_code=422,
        )

    if len(new_password) < 8:
        return _render_account_page(
            request,
            user,
            error_text="Новый пароль должен быть не короче 8 символов.",
            status_code=422,
        )

    if new_password != new_password_confirm:
        return _render_account_page(
            request,
            user,
            error_text="Подтверждение нового пароля не совпадает.",
            status_code=422,
        )

    if verify_password(new_password, user.password_hash):
        return _render_account_page(
            request,
            user,
            error_text="Новый пароль должен отличаться от текущего.",
            status_code=422,
        )

    user.password_hash = hash_password(new_password)
    db.commit()

    return _render_account_page(
        request,
        user,
        success_text="Пароль успешно изменен.",
    )


@app.get("/", response_class=HTMLResponse)
def index_page(
    request: Request,
    db: Session = Depends(get_db),
):
    user = _current_user_from_session(request, db)
    if not user:
        return _redirect_to_login()
    if not _can_manage_registry(user):
        return RedirectResponse(url="/active", status_code=303)

    context = _template_context(request, user)
    context.update(
        {
            "field_help": FIELD_HELP,
            "api_base": "/api",
            "ws_path": "/ws/registry",
        }
    )
    return templates.TemplateResponse("index.html", context)


@app.get("/active", response_class=HTMLResponse)
def active_page(
    request: Request,
    db: Session = Depends(get_db),
):
    user = _current_user_from_session(request, db)
    if not user:
        context = {
            "request": request,
            "current_user": None,
            "is_authenticated": False,
            "is_admin": False,
            "is_editor": False,
            "is_client": True,
            "can_manage_registry": False,
            "role_label": "Гость",
            "api_base": "/api/public",
            "ws_path": "/ws/registry",
            "can_edit": False,
            "can_full_edit": False,
            "can_delete": False,
            "user_role": "public",
            "export_url": "/api/public/vacations/export?tab=current",
        }
        return templates.TemplateResponse("active.html", context)

    context = _template_context(request, user)
    context.update(
        {
            "is_authenticated": True,
            "api_base": "/api",
            "ws_path": "/ws/registry",
            "can_edit": _can_manage_registry(user),
            "can_full_edit": _can_manage_registry(user),
            "can_delete": _is_admin(user),
            "user_role": user.role.value,
            "export_url": "/api/vacations/active/export",
        }
    )
    return templates.TemplateResponse("active.html", context)


@app.get("/public", response_class=HTMLResponse)
def public_page(request: Request):
    return templates.TemplateResponse(
        "public.html",
        {
            "request": request,
            "api_base": "/api/public",
            "ws_path": "/ws/registry",
        },
    )


@app.get("/plans", response_class=HTMLResponse)
def plans_page(
    request: Request,
    db: Session = Depends(get_db),
):
    user = _current_user_from_session(request, db)
    if not user:
        return _redirect_to_login()
    if not _is_admin(user):
        return RedirectResponse(url="/active", status_code=303)

    context = _template_context(request, user)
    context.update(
        {
            "api_base": "/api",
            "ws_path": "/ws/registry",
        }
    )
    return templates.TemplateResponse("plans.html", context)


@app.get("/admin", response_class=HTMLResponse)
def admin_page(
    request: Request,
    db: Session = Depends(get_db),
):
    user = _current_user_from_session(request, db)
    if not user:
        return _redirect_to_login()
    if not _is_admin(user):
        return RedirectResponse(url="/active", status_code=303)
    return _render_admin_page(request, db, user)


@app.post("/admin/users", response_class=HTMLResponse)
def create_user_from_admin_page(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    role: str = Form(default="client"),
    db: Session = Depends(get_db),
):
    current_user = _current_user_from_session(request, db)
    if not current_user:
        return _redirect_to_login()
    if not _is_admin(current_user):
        return RedirectResponse(url="/active", status_code=303)

    normalized_username = normalize_username(username)

    if len(normalized_username) < 3:
        return _render_admin_page(
            request,
            db,
            current_user,
            error_text="Логин должен быть не короче 3 символов.",
            status_code=422,
        )

    if len(password) < 8:
        return _render_admin_page(
            request,
            db,
            current_user,
            error_text="Пароль должен быть не короче 8 символов.",
            status_code=422,
        )

    allowed_roles = {UserRole.ADMIN.value, UserRole.EDITOR.value, UserRole.CLIENT.value, UserRole.VIEWER.value}
    role_value = role if role in allowed_roles else UserRole.CLIENT.value
    if role_value == UserRole.VIEWER.value:
        role_value = UserRole.CLIENT.value
    exists = db.scalar(select(UserAccount.id).where(UserAccount.username == normalized_username))
    if exists:
        return _render_admin_page(
            request,
            db,
            current_user,
            error_text="Пользователь с таким логином уже существует.",
            status_code=409,
        )

    new_user = UserAccount(
        username=normalized_username,
        password_hash=hash_password(password),
        role=UserRole(role_value),
        is_active=True,
    )
    db.add(new_user)
    db.commit()

    return _render_admin_page(
        request,
        db,
        current_user,
        success_text="Пользователь создан.",
        status_code=201,
    )


@app.post("/admin/users/{user_id}/password", response_class=HTMLResponse)
def reset_user_password_from_admin_page(
    user_id: int,
    request: Request,
    new_password: str = Form(...),
    db: Session = Depends(get_db),
):
    current_user = _current_user_from_session(request, db)
    if not current_user:
        return _redirect_to_login()
    if not _is_admin(current_user):
        return RedirectResponse(url="/active", status_code=303)

    target_user = db.get(UserAccount, user_id)
    if not target_user:
        return _render_admin_page(
            request,
            db,
            current_user,
            error_text="Пользователь не найден.",
            status_code=404,
        )

    if len(new_password) < 8:
        return _render_admin_page(
            request,
            db,
            current_user,
            error_text="Новый пароль должен быть не короче 8 символов.",
            status_code=422,
        )

    target_user.password_hash = hash_password(new_password)
    db.commit()

    return _render_admin_page(
        request,
        db,
        current_user,
        success_text=f"Пароль пользователя '{target_user.username}' обновлен.",
    )


@app.post("/admin/notification-config", response_class=HTMLResponse)
def update_notification_config_from_admin_page(
    request: Request,
    repeat_interval_minutes: int = Form(default=0),
    db: Session = Depends(get_db),
):
    current_user = _current_user_from_session(request, db)
    if not current_user:
        return _redirect_to_login()
    if not _is_admin(current_user):
        return RedirectResponse(url="/active", status_code=303)

    if repeat_interval_minutes < MIN_NOTIFICATION_REPEAT_MINUTES or repeat_interval_minutes > MAX_NOTIFICATION_REPEAT_MINUTES:
        return _render_admin_page(
            request,
            db,
            current_user,
            error_text=(
                f"Интервал повтора должен быть от {MIN_NOTIFICATION_REPEAT_MINUTES} "
                f"до {MAX_NOTIFICATION_REPEAT_MINUTES} минут."
            ),
            status_code=422,
        )

    config = _ensure_notification_config(db)
    config.repeat_interval_minutes = _normalize_repeat_interval_minutes(repeat_interval_minutes, 0)
    db.commit()

    return _render_admin_page(
        request,
        db,
        current_user,
        success_text=f"Интервал повтора уведомлений сохранен: {config.repeat_interval_minutes} мин.",
    )


@app.post("/admin/client-update-config", response_class=HTMLResponse)
def update_client_update_config_from_admin_page(
    request: Request,
    enabled: str = Form(default="1"),
    latest_version: str = Form(default=""),
    windows_installer_url: str = Form(default=""),
    linux_deb_url: str = Form(default=""),
    linux_rpm_url: str = Form(default=""),
    check_interval_minutes: int = Form(default=60),
    db: Session = Depends(get_db),
):
    current_user = _current_user_from_session(request, db)
    if not current_user:
        return _redirect_to_login()
    if not _is_admin(current_user):
        return RedirectResponse(url="/active", status_code=303)

    normalized_version = _normalize_update_version(latest_version, "")
    if not normalized_version:
        return _render_admin_page(
            request,
            db,
            current_user,
            error_text="Укажите версию клиента (например, 1.2.0).",
            status_code=422,
        )

    update_config = _ensure_client_update_config(db)
    update_config.enabled = str(enabled).strip().lower() in {"1", "true", "yes", "on"}
    update_config.latest_version = normalized_version
    update_config.windows_installer_url = _normalize_text(windows_installer_url) or None
    update_config.linux_deb_url = _normalize_text(linux_deb_url) or None
    update_config.linux_rpm_url = _normalize_text(linux_rpm_url) or None
    update_config.check_interval_minutes = _normalize_update_check_interval_minutes(check_interval_minutes, 60)
    db.commit()

    return _render_admin_page(
        request,
        db,
        current_user,
        success_text=f"Настройки автообновления сохранены. Актуальная версия: {update_config.latest_version}.",
    )


def _serialize_sync_log(log: RegistrySyncLog) -> dict:
    return {
        "id": log.id,
        "action_type": log.action_type,
        "source_name": log.source_name,
        "actor_username": log.actor_username,
        "payload": log.payload or {},
        "created_at": log.created_at.isoformat() if log.created_at else None,
    }


def _resolve_client_ip(request: Request, payload_ip: str | None = None) -> str:
    xff = request.headers.get("x-forwarded-for", "")
    if xff:
        first = xff.split(",")[0].strip()
        if first:
            return first
    real_ip = request.headers.get("x-real-ip", "").strip()
    if real_ip:
        return real_ip
    if request.client and request.client.host:
        return str(request.client.host)
    return _normalize_text(payload_ip) or ""


def _normalize_mac_address(value: str | None) -> str:
    raw = "".join(ch for ch in str(value or "").lower() if ch in "0123456789abcdef")
    if len(raw) != 12 or raw in {"000000000000", "ffffffffffff"}:
        return ""
    try:
        first_octet = int(raw[:2], 16)
    except ValueError:
        return ""
    if first_octet & 1:
        return ""
    return ":".join(raw[index : index + 2] for index in range(0, 12, 2))


def _normalize_client_hostname(value: str | None) -> str:
    hostname = _normalize_text(value).lower().strip(".")
    if not hostname or hostname in {"localhost", "localhost.localdomain", "unknown-host"}:
        return ""
    # Linux clients often report FQDN values like pc01224-1.localdomain.
    # For device identity we only need the stable host name, not the domain.
    return hostname.split(".", 1)[0]


def _upsert_client_heartbeat(db: Session, request: Request, payload: ClientHeartbeatUpsert) -> ClientHeartbeat | None:
    ip_address = _resolve_client_ip(request, payload.ip_address)
    mac_address = _normalize_mac_address(payload.mac_address) or _mac_from_client_id(payload.client_id)
    normalized_hostname = _normalize_client_hostname(payload.hostname)
    now_utc = datetime.now(timezone.utc)
    heartbeat = None

    if not mac_address:
        no_mac_filter = and_(
            or_(ClientHeartbeat.mac_address.is_(None), ClientHeartbeat.mac_address == ""),
            or_(ClientHeartbeat.client_id.is_(None), ~func.lower(ClientHeartbeat.client_id).like("mac-%")),
        )
        legacy_filters = [ClientHeartbeat.client_id == payload.client_id]
        if normalized_hostname:
            legacy_filters.append(
                and_(
                    no_mac_filter,
                    or_(
                        func.lower(ClientHeartbeat.hostname) == normalized_hostname,
                        func.lower(ClientHeartbeat.hostname).like(f"{normalized_hostname}.%"),
                    ),
                )
            )
        if ip_address:
            legacy_filters.append(and_(no_mac_filter, ClientHeartbeat.ip_address == ip_address))

        legacy_rows = db.scalars(select(ClientHeartbeat).where(or_(*legacy_filters))).all()
        for legacy_row in legacy_rows:
            legacy_mac = _normalize_mac_address(legacy_row.mac_address) or _mac_from_client_id(legacy_row.client_id)
            if not legacy_mac:
                db.delete(legacy_row)
        db.commit()
        return None

    if mac_address:
        heartbeat = db.scalar(select(ClientHeartbeat).where(ClientHeartbeat.mac_address == mac_address))
    if not heartbeat and normalized_hostname:
        heartbeat = db.scalar(
            select(ClientHeartbeat)
            .where(
                or_(
                    func.lower(ClientHeartbeat.hostname) == normalized_hostname,
                    func.lower(ClientHeartbeat.hostname).like(f"{normalized_hostname}.%"),
                )
            )
            .order_by(ClientHeartbeat.last_seen_at.desc(), ClientHeartbeat.id.desc())
        )
    if not heartbeat and ip_address:
        heartbeat = db.scalar(
            select(ClientHeartbeat)
            .where(ClientHeartbeat.ip_address == ip_address)
            .order_by(ClientHeartbeat.last_seen_at.desc(), ClientHeartbeat.id.desc())
        )
    if not heartbeat:
        heartbeat = db.scalar(select(ClientHeartbeat).where(ClientHeartbeat.client_id == payload.client_id))

    if not heartbeat:
        heartbeat = ClientHeartbeat(
            client_id=payload.client_id,
            created_at=now_utc,
        )
        db.add(heartbeat)
        db.flush()
    elif heartbeat.client_id != payload.client_id:
        conflicting = db.scalar(
            select(ClientHeartbeat).where(
                ClientHeartbeat.client_id == payload.client_id,
                ClientHeartbeat.id != heartbeat.id,
            )
        )
        if conflicting:
            db.delete(conflicting)
            db.flush()
        heartbeat.client_id = payload.client_id

    heartbeat.hostname = _normalize_text(payload.hostname)
    heartbeat.username = _normalize_text(payload.username)
    heartbeat.os_name = _normalize_text(payload.os_name)
    heartbeat.os_version = _normalize_text(payload.os_version)
    heartbeat.mac_address = mac_address or None
    heartbeat.app_version = _normalize_text(payload.app_version)
    heartbeat.app_channel = _normalize_text(payload.app_channel) or "stable"
    heartbeat.ip_address = ip_address or None
    heartbeat.mode = _normalize_text(payload.mode)
    heartbeat.last_notification_id = payload.last_notification_id if isinstance(payload.last_notification_id, int) else None
    heartbeat.update_supported = bool(payload.update_supported)
    heartbeat.update_status = _normalize_text(payload.update_status)
    heartbeat.update_error = _normalize_text(payload.update_error)
    heartbeat.last_seen_at = now_utc

    duplicate_filters = []
    if mac_address:
        duplicate_filters.append(ClientHeartbeat.mac_address == mac_address)
    if normalized_hostname:
        duplicate_filters.append(
            and_(
                or_(ClientHeartbeat.mac_address.is_(None), ClientHeartbeat.mac_address == ""),
                or_(
                    func.lower(ClientHeartbeat.hostname) == normalized_hostname,
                    func.lower(ClientHeartbeat.hostname).like(f"{normalized_hostname}.%"),
                ),
            )
        )
    if ip_address:
        duplicate_filters.append(
            and_(
                or_(ClientHeartbeat.mac_address.is_(None), ClientHeartbeat.mac_address == ""),
                ClientHeartbeat.ip_address == ip_address,
            )
        )
    if duplicate_filters:
        duplicates = db.scalars(
            select(ClientHeartbeat).where(
                ClientHeartbeat.id != heartbeat.id,
                or_(*duplicate_filters),
            )
        ).all()
        for duplicate in duplicates:
            db.delete(duplicate)

    db.commit()
    db.refresh(heartbeat)
    return heartbeat


def _normalize_dt_utc(value: datetime | None) -> datetime | None:
    if value is None:
        return None
    if value.tzinfo is None:
        return value.replace(tzinfo=timezone.utc)
    return value.astimezone(timezone.utc)


def _mac_from_client_id(value: str | None) -> str:
    text_value = _normalize_text(value).lower()
    if not text_value.startswith("mac-"):
        return ""
    return _normalize_mac_address(text_value.removeprefix("mac-"))


def _client_heartbeat_identity(row: ClientHeartbeat) -> str:
    mac_address = _normalize_mac_address(row.mac_address) or _mac_from_client_id(row.client_id)
    if mac_address:
        return f"mac:{mac_address}"
    hostname = _normalize_client_hostname(row.hostname)
    if hostname:
        return f"host:{hostname}"
    ip_address = _normalize_text(row.ip_address)
    if ip_address:
        return f"ip:{ip_address}"
    return f"id:{row.client_id}"


def _cleanup_client_heartbeat_duplicates(db: Session) -> int:
    rows = db.scalars(
        select(ClientHeartbeat).order_by(ClientHeartbeat.last_seen_at.desc(), ClientHeartbeat.id.desc())
    ).all()
    seen: set[str] = set()
    seen_ip_without_mac: set[str] = set()
    deleted_count = 0
    changed = False

    for row in rows:
        mac_address = _normalize_mac_address(row.mac_address) or _mac_from_client_id(row.client_id)
        if not mac_address:
            db.delete(row)
            deleted_count += 1
            changed = True
            continue
        if mac_address and row.mac_address != mac_address:
            row.mac_address = mac_address
            changed = True
        identity = _client_heartbeat_identity(row)
        if identity in seen:
            db.delete(row)
            deleted_count += 1
            changed = True
            continue
        seen.add(identity)
        ip_address = _normalize_text(row.ip_address)
        if not mac_address and ip_address:
            if ip_address in seen_ip_without_mac:
                db.delete(row)
                deleted_count += 1
                changed = True
                continue
            seen_ip_without_mac.add(ip_address)

    if changed:
        db.commit()
    return deleted_count


def _ensure_client_heartbeat_unique_indexes() -> None:
    where_clause = "mac_address IS NOT NULL AND mac_address <> ''"
    with engine.begin() as conn:
        try:
            conn.execute(
                text(
                    "CREATE UNIQUE INDEX IF NOT EXISTS uq_client_heartbeats_mac_address "
                    f"ON client_heartbeats (mac_address) WHERE {where_clause}"
                )
            )
        except Exception:
            pass


def _build_client_monitor_rows(db: Session, *, latest_version: str) -> list[ClientHeartbeatMonitorResponse]:
    now_utc = datetime.now(timezone.utc)
    online_delta_seconds = max(30, CLIENT_ONLINE_WINDOW_SECONDS)
    _cleanup_client_heartbeat_duplicates(db)
    rows = db.scalars(select(ClientHeartbeat).order_by(ClientHeartbeat.last_seen_at.desc(), ClientHeartbeat.id.desc())).all()

    result: list[ClientHeartbeatMonitorResponse] = []
    seen_fingerprints: set[str] = set()
    for row in rows:
        mac_address = _normalize_mac_address(row.mac_address) or _mac_from_client_id(row.client_id)
        if not mac_address:
            continue
        fingerprint = _client_heartbeat_identity(row)
        if fingerprint in seen_fingerprints:
            continue
        seen_fingerprints.add(fingerprint)

        last_seen_utc = _normalize_dt_utc(row.last_seen_at) or now_utc
        age_seconds = max(0.0, (now_utc - last_seen_utc).total_seconds())
        app_version = _normalize_text(row.app_version)
        result.append(
            ClientHeartbeatMonitorResponse(
                client_id=row.client_id,
                hostname=row.hostname,
                username=row.username,
                os_name=row.os_name,
                os_version=row.os_version,
                mac_address=mac_address or row.mac_address,
                app_version=app_version or "unknown",
                app_channel=row.app_channel,
                ip_address=row.ip_address,
                mode=row.mode,
                last_notification_id=row.last_notification_id,
                update_supported=bool(row.update_supported),
                update_status=row.update_status,
                update_error=row.update_error,
                last_seen_at=last_seen_utc,
                is_online=age_seconds <= online_delta_seconds,
                is_outdated=_version_is_newer(latest_version, app_version),
                latest_version=latest_version,
            )
        )
    return result


def _build_return_messages(db: Session, record_ids: list[int]) -> list[str]:
    if not record_ids:
        return []

    rows = db.scalars(
        select(VacationRecord)
        .where(VacationRecord.id.in_(record_ids))
        .order_by(VacationRecord.employee_full_name.asc(), VacationRecord.end_date.asc())
    ).all()

    return [
        (
            f"{row.employee_full_name} вышел из отпуска "
            f"({row.end_date.isoformat()}), запись скрыта из клиентской таблицы."
        )
        for row in rows
    ]


@app.get("/api/health")
def health_check():
    return {"status": "ok", "time": datetime.now(timezone.utc).isoformat()}


@app.get("/api/me")
def me(user: UserAccount = Depends(require_user)):
    return {
        "id": user.id,
        "username": user.username,
        "role": user.role.value,
        "is_admin": user.role == UserRole.ADMIN,
        "is_editor": user.role == UserRole.EDITOR,
        "is_client": user.role in {UserRole.CLIENT, UserRole.VIEWER},
        "can_manage_registry": user.role in {UserRole.ADMIN, UserRole.EDITOR},
    }


@app.get("/api/notification-config", response_model=NotificationConfigResponse)
def get_notification_config(db: Session = Depends(get_db)):
    return NotificationConfigResponse(
        repeat_interval_minutes=_get_notification_repeat_interval_minutes(db),
    )


@app.get("/api/client-update-config", response_model=ClientUpdateConfigResponse)
def get_client_update_config(_: UserAccount = Depends(require_admin), db: Session = Depends(get_db)):
    config = _ensure_client_update_config(db)
    return _serialize_client_update_config(config)


@app.put("/api/client-update-config", response_model=ClientUpdateConfigResponse)
def update_client_update_config(
    payload: ClientUpdateConfigUpdate,
    _: UserAccount = Depends(require_admin),
    db: Session = Depends(get_db),
):
    config = _ensure_client_update_config(db)
    config.enabled = bool(payload.enabled)
    config.latest_version = _normalize_update_version(payload.latest_version, DEFAULT_CLIENT_VERSION)
    config.windows_installer_url = _normalize_text(payload.windows_installer_url) or None
    config.linux_deb_url = _normalize_text(payload.linux_deb_url) or None
    config.linux_rpm_url = _normalize_text(payload.linux_rpm_url) or None
    config.check_interval_minutes = _normalize_update_check_interval_minutes(payload.check_interval_minutes, 60)
    db.commit()
    db.refresh(config)
    return _serialize_client_update_config(config)


@app.get("/api/client-heartbeats", response_model=list[ClientHeartbeatMonitorResponse])
def list_client_heartbeats(_: UserAccount = Depends(require_admin), db: Session = Depends(get_db)):
    update_config = _ensure_client_update_config(db)
    latest_version = _normalize_update_version(update_config.latest_version, DEFAULT_CLIENT_VERSION)
    return _build_client_monitor_rows(db, latest_version=latest_version)


@app.post("/api/client-heartbeat", response_model=ClientHeartbeatResponse)
def upsert_client_heartbeat(
    payload: ClientHeartbeatUpsert,
    request: Request,
    db: Session = Depends(get_db),
):
    _require_client_heartbeat_token(request)
    row = _upsert_client_heartbeat(db, request, payload)
    return ClientHeartbeatResponse(
        status="ok",
        client_id=row.client_id if row else payload.client_id,
        server_time=datetime.now(timezone.utc),
    )


@app.get("/api/client-update-manifest", response_model=ClientUpdateManifestResponse)
def get_client_update_manifest(
    request: Request,
    platform: str = Query(default="windows", description="windows | linux | linux-rpm | linux-deb"),
    db: Session = Depends(get_db),
):
    config = _ensure_client_update_config(db)
    normalized_platform = _normalize_text(platform).lower() or "windows"
    default_urls = _build_default_client_update_urls(request)
    update_url = _resolve_client_update_url(
        config,
        normalized_platform,
        fallback_windows_url=default_urls["windows"],
        fallback_linux_deb_url=default_urls["linux-deb"],
        fallback_linux_rpm_url=default_urls["linux-rpm"],
    )
    windows_installer_url = _normalize_text(config.windows_installer_url) or default_urls["windows"]
    linux_deb_url = _normalize_text(config.linux_deb_url) or default_urls["linux-deb"]
    linux_rpm_url = _normalize_text(config.linux_rpm_url) or default_urls["linux-rpm"]
    latest_version = _normalize_update_version(config.latest_version, DEFAULT_CLIENT_VERSION)
    return ClientUpdateManifestResponse(
        enabled=bool(config.enabled),
        latest_version=latest_version,
        check_interval_minutes=_normalize_update_check_interval_minutes(config.check_interval_minutes, 60),
        update_url=update_url,
        windows_installer_url=windows_installer_url or None,
        linux_deb_url=linux_deb_url or None,
        linux_rpm_url=linux_rpm_url or None,
        generated_at=datetime.now(timezone.utc),
    )


@app.get("/api/field-help")
def field_help(_: UserAccount = Depends(require_admin)):
    return JSONResponse(FIELD_HELP)


@app.get("/api/users", response_model=list[UserResponse])
def list_users(_: UserAccount = Depends(require_admin), db: Session = Depends(get_db)):
    stmt = select(UserAccount).order_by(UserAccount.created_at.asc(), UserAccount.id.asc())
    return db.scalars(stmt).all()


@app.post("/api/users", response_model=UserResponse, status_code=201)
def create_user(payload: UserCreate, _: UserAccount = Depends(require_admin), db: Session = Depends(get_db)):
    exists = db.scalar(select(UserAccount.id).where(UserAccount.username == payload.username))
    if exists:
        raise HTTPException(status_code=409, detail="Пользователь с таким логином уже существует")

    new_user = UserAccount(
        username=payload.username,
        password_hash=hash_password(payload.password),
        role=payload.role,
        is_active=True,
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    return new_user


@app.get("/api/notifications", response_model=list[NotificationEventResponse])
def list_notifications(
    after_id: int = Query(default=0, ge=0, description="Вернуть события с id больше этого значения"),
    limit: int = Query(default=100, ge=1, le=1000, description="Лимит количества событий"),
    db: Session = Depends(get_db),
):
    repeat_interval_minutes = _get_notification_repeat_interval_minutes(db)
    stmt = (
        select(NotificationEvent)
        .where(NotificationEvent.id > after_id)
        .order_by(NotificationEvent.id.asc())
        .limit(limit)
    )
    events = db.scalars(stmt).all()
    return [_serialize_notification(event, repeat_interval_minutes) for event in events]


def _build_plans_statement(
    *,
    employee_full_name: str = "",
    service: str = "",
    period_from: date | None = None,
    period_to: date | None = None,
    include_finished: bool = True,
):
    if period_from and period_to and period_to < period_from:
        raise HTTPException(status_code=422, detail="Дата окончания периода фильтра не может быть раньше даты начала.")

    stmt = select(VacationRecord)

    if employee_full_name.strip():
        term = _normalize_text(employee_full_name).lower()
        stmt = stmt.where(func.lower(VacationRecord.employee_full_name).contains(term))

    if service.strip():
        term = _normalize_text(service).lower()
        stmt = stmt.where(func.lower(VacationRecord.service).contains(term))

    if period_from:
        stmt = stmt.where(VacationRecord.end_date >= period_from)
    if period_to:
        stmt = stmt.where(VacationRecord.start_date <= period_to)

    if not include_finished:
        stmt = stmt.where(VacationRecord.status.in_(ACTIVE_STATUS_DB_VALUES))

    return stmt.order_by(VacationRecord.start_date.asc(), VacationRecord.employee_full_name.asc())


@app.get("/api/plans", response_model=list[VacationResponse])
def list_plans(
    employee_full_name: str = Query(default="", description="Фильтр по ФИО"),
    service: str = Query(default="", description="Фильтр по услуге"),
    period_from: date | None = Query(default=None, description="Период: дата от"),
    period_to: date | None = Query(default=None, description="Период: дата до"),
    include_finished: bool = Query(default=True, description="Показывать завершенные записи"),
    _: UserAccount = Depends(require_admin),
    db: Session = Depends(get_db),
):
    stmt = _build_plans_statement(
        employee_full_name=employee_full_name,
        service=service,
        period_from=period_from,
        period_to=period_to,
        include_finished=include_finished,
    )
    return db.scalars(stmt).all()


@app.get("/api/plans/export")
def export_plans(
    employee_full_name: str = Query(default="", description="Фильтр по ФИО"),
    service: str = Query(default="", description="Фильтр по услуге"),
    period_from: date | None = Query(default=None, description="Период: дата от"),
    period_to: date | None = Query(default=None, description="Период: дата до"),
    include_finished: bool = Query(default=True, description="Показывать завершенные записи"),
    _: UserAccount = Depends(require_admin),
    db: Session = Depends(get_db),
):
    stmt = _build_plans_statement(
        employee_full_name=employee_full_name,
        service=service,
        period_from=period_from,
        period_to=period_to,
        include_finished=include_finished,
    )
    rows = db.scalars(stmt).all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "План отпусков"
    sheet.append(
        [
            "ID",
            "ФИО сотрудника",
            "Особый сотрудник",
            "Должность/Должности",
            "Статус",
            "Услуга",
            "Заместитель/и (ФИО)",
            "Фактическая должность сотрудника (замещающего)",
            "Дата начала отпуска",
            "Дата окончания отпуска",
            "Памятка",
        ]
    )

    for row in rows:
        sheet.append(
            [
                row.id,
                row.employee_full_name,
                "Да" if row.is_special_employee else "Нет",
                "\n".join(row.employee_positions),
                row.status.value,
                row.service,
                "\n".join(dep.get("deputy_full_name", "") for dep in row.deputies or []),
                "\n".join(dep.get("deputy_actual_position", "") for dep in row.deputies or []),
                row.start_date.isoformat(),
                row.end_date.isoformat(),
                row.memo or "",
            ]
        )

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    filename = f"vacations_plan_{datetime.now(APP_TZ).date().isoformat()}.xlsx"
    headers = {"Content-Disposition": f"attachment; filename={filename}"}
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.get("/api/sync-logs")
def list_sync_logs(
    limit: int = Query(default=50, ge=1, le=500),
    _: UserAccount = Depends(require_admin),
    db: Session = Depends(get_db),
):
    stmt = (
        select(RegistrySyncLog)
        .order_by(RegistrySyncLog.created_at.desc(), RegistrySyncLog.id.desc())
        .limit(limit)
    )
    logs = db.scalars(stmt).all()
    return [_serialize_sync_log(log) for log in logs]


@app.post("/api/maintenance/apply-today")
async def apply_plan_today(
    current_user: UserAccount = Depends(require_admin),
    db: Session = Depends(get_db),
):
    maintenance = run_registry_maintenance(db, datetime.now(APP_TZ).date(), retention_days=365)
    changed_ids = maintenance.changed_ids
    activated_ids = sorted(set(maintenance.activated_ids))
    # Notify users only when a record appears in current vacations table.
    should_notify = bool(activated_ids)

    _add_sync_log(
        db,
        action_type="manual_sync",
        actor_username=current_user.username,
        payload={
            "activated_count": len(maintenance.activated_ids),
            "finalized_count": len(maintenance.finalized_ids),
            "deleted_count": len(maintenance.deleted_ids),
            "changed_ids": changed_ids,
            "activated_ids": activated_ids,
        },
    )
    db.commit()

    if should_notify:
        event, special_names = _create_registry_notification_for_changes(db, activated_ids)
        await _broadcast_registry_update(activated_ids, event, special_employee_names=special_names)

    return {
        "activated_count": len(maintenance.activated_ids),
        "finalized_count": len(maintenance.finalized_ids),
        "deleted_count": len(maintenance.deleted_ids),
        "changed_ids": changed_ids,
        "activated_ids": activated_ids,
    }


@app.post("/api/vacations", response_model=VacationResponse, status_code=201)
async def create_vacation(
    payload: VacationCreate,
    _: UserAccount = Depends(require_registry_editor),
    db: Session = Depends(get_db),
):
    _validate_deputies(payload)
    duplicate = _find_duplicate_vacation_record(db, payload.employee_full_name, payload.start_date, payload.end_date)
    if duplicate:
        raise HTTPException(
            status_code=409,
            detail=(
                f"Дублирующая запись уже существует (ID {duplicate.id}): "
                f"{duplicate.employee_full_name}, {duplicate.start_date} - {duplicate.end_date}."
            ),
        )

    record = VacationRecord(
        employee_full_name=payload.employee_full_name,
        is_special_employee=payload.is_special_employee,
        employee_positions=payload.employee_positions,
        status=payload.status,
        service=payload.service,
        deputies=[deputy.model_dump() for deputy in payload.deputies],
        memo=payload.memo,
        start_date=payload.start_date,
        end_date=payload.end_date,
    )

    _apply_status_by_dates(record)
    db.add(record)
    db.commit()
    db.refresh(record)

    today = datetime.now(APP_TZ).date()
    if _is_record_active_today(record, today):
        event, special_names = _create_registry_notification_for_changes(db, [record.id])
        await _broadcast_registry_update([record.id], event, special_employee_names=special_names)
    return record


@app.put("/api/vacations/{record_id}", response_model=VacationResponse)
async def update_vacation(
    record_id: int,
    payload: VacationUpdate,
    _: UserAccount = Depends(require_registry_editor),
    db: Session = Depends(get_db),
):
    record = db.get(VacationRecord, record_id)
    if not record:
        raise HTTPException(status_code=404, detail="Запись не найдена")

    today = datetime.now(APP_TZ).date()
    was_active_before = _is_record_active_today(record, today)

    _validate_deputies(payload)

    duplicate = _find_duplicate_vacation_record(
        db,
        payload.employee_full_name,
        payload.start_date,
        payload.end_date,
        exclude_id=record_id,
    )
    if duplicate:
        raise HTTPException(
            status_code=409,
            detail=f"Нельзя сохранить дубликат. Запись с такими ФИО и периодом уже есть (ID {duplicate.id}).",
        )

    record.employee_full_name = payload.employee_full_name
    record.is_special_employee = payload.is_special_employee
    record.employee_positions = payload.employee_positions
    record.status = payload.status
    record.service = payload.service
    record.deputies = [deputy.model_dump() for deputy in payload.deputies]
    record.memo = payload.memo
    record.start_date = payload.start_date
    record.end_date = payload.end_date

    _apply_status_by_dates(record)

    db.commit()
    db.refresh(record)

    is_active_now = _is_record_active_today(record, today)
    if (not was_active_before) and is_active_now:
        event, special_names = _create_registry_notification_for_changes(db, [record.id])
        await _broadcast_registry_update([record.id], event, special_employee_names=special_names)
    return record

@app.post("/api/vacations/import")
async def import_vacations_excel(
    file: UploadFile = File(...),
    current_user: UserAccount = Depends(require_registry_editor),
    db: Session = Depends(get_db),
):
    filename = file.filename or ""
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=422, detail="Поддерживается только формат .xlsx")

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=422, detail="Файл пустой.")

    try:
        workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Не удалось прочитать Excel-файл: {exc}") from exc

    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=422, detail="В файле отсутствуют данные.")

    try:
        header_row_num, column_indexes = _find_excel_header_row(rows)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc

    today = datetime.now(APP_TZ).date()
    now_utc = datetime.now(timezone.utc)

    def get_cell(row: tuple[object, ...], key: str) -> object | None:
        idx = column_indexes.get(key)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    parsed_rows: list[dict] = []
    imported_keys_by_employee: dict[str, set[tuple[str, str, str, str]]] = {}
    seen_import_keys: set[tuple[str, str, str, str]] = set()

    created_ids: list[int] = []
    updated_count = 0
    reconciled_archived_count = 0
    reconciled_deleted_count = 0
    duplicate_count = 0
    skipped_empty_rows = 0
    errors: list[str] = []

    for row_num, row in enumerate(rows[header_row_num:], start=header_row_num + 1):
        if row is None:
            skipped_empty_rows += 1
            continue

        if all(_normalize_text(str(cell) if cell is not None else "") == "" for cell in row):
            skipped_empty_rows += 1
            continue

        raw_start = get_cell(row, "start_date")
        raw_end = get_cell(row, "end_date")
        if (
            _normalize_text(str(raw_start) if raw_start is not None else "") == ""
            and _normalize_text(str(raw_end) if raw_end is not None else "") == ""
        ):
            skipped_empty_rows += 1
            continue

        try:
            employee_full_name = _normalize_text(str(get_cell(row, "employee_full_name") or ""))
            if not employee_full_name:
                raise ValueError("Не заполнено поле 'ФИО сотрудника'.")

            employee_positions = _split_cell_lines(get_cell(row, "employee_positions"))
            if not employee_positions:
                raise ValueError("Не заполнено поле 'Должность/Должности'.")

            is_special_employee = _parse_excel_special_employee(get_cell(row, "is_special_employee"))

            service = _normalize_text(str(get_cell(row, "service") or ""))
            if not service:
                service = "Отпуск"

            status = _parse_excel_status(get_cell(row, "status"))
            start_date = _parse_excel_date(
                get_cell(row, "start_date"),
                row_num=row_num,
                column_name="Дата начала отпуска",
            )
            end_date = _parse_excel_date(
                get_cell(row, "end_date"),
                row_num=row_num,
                column_name="Дата окончания отпуска",
            )
            if end_date < start_date:
                raise ValueError("Дата окончания отпуска не может быть раньше даты начала.")

            deputy_full_names = _split_cell_lines(get_cell(row, "deputy_full_names"))
            deputy_actual_positions = _split_cell_lines(get_cell(row, "deputy_actual_positions"))
            if deputy_full_names:
                if not deputy_actual_positions:
                    deputy_actual_positions = ["Не назначен"] * len(deputy_full_names)
                deputies = _build_deputies_from_columns(employee_positions, deputy_full_names, deputy_actual_positions)
            else:
                deputies = [
                    {
                        "vacation_position": position,
                        "deputy_full_name": "Не назначен",
                        "deputy_actual_position": "Не назначен",
                    }
                    for position in employee_positions
                ]

            memo_value = _normalize_text(str(get_cell(row, "memo") or ""))
            payload = VacationCreate(
                employee_full_name=employee_full_name,
                is_special_employee=is_special_employee,
                employee_positions=employee_positions,
                status=status,
                service=service,
                deputies=deputies,
                memo=memo_value or None,
                start_date=start_date,
                end_date=end_date,
            )
            _validate_deputies(payload)

            import_key = _vacation_identity_key(
                payload.employee_full_name,
                payload.start_date,
                payload.end_date,
                service=payload.service,
            )
            if import_key in seen_import_keys:
                duplicate_count += 1
                continue

            parsed_rows.append(
                {
                    "payload": payload,
                    "imported_deputies": [deputy.model_dump() for deputy in payload.deputies],
                    "deputies_provided": bool(deputy_full_names),
                    "memo_provided": bool(memo_value),
                }
            )
            seen_import_keys.add(import_key)
            imported_keys_by_employee.setdefault(import_key[0], set()).add(import_key)
        except Exception as exc:
            errors.append(f"Строка {row_num}: {exc}")

    activated_ids: set[int] = set()
    all_changed_ids: set[int] = set()

    for row_data in parsed_rows:
        payload: VacationCreate = row_data["payload"]
        imported_deputies: list[dict] = row_data["imported_deputies"]
        deputies_provided: bool = row_data["deputies_provided"]
        memo_provided: bool = row_data["memo_provided"]

        existing = _find_duplicate_vacation_record(
            db,
            payload.employee_full_name,
            payload.start_date,
            payload.end_date,
        )

        if existing:
            was_active = _is_record_active_today(existing, today)
            all_changed_ids.add(existing.id)

            existing.employee_full_name = payload.employee_full_name
            existing.is_special_employee = payload.is_special_employee
            existing.employee_positions = payload.employee_positions
            existing.status = payload.status
            existing.service = payload.service
            existing.start_date = payload.start_date
            existing.end_date = payload.end_date

            if deputies_provided or not existing.deputies or _deputies_are_placeholders(existing.deputies):
                existing.deputies = imported_deputies

            if memo_provided:
                existing.memo = payload.memo
            elif existing.memo is None:
                existing.memo = payload.memo

            _apply_status_by_dates(existing)
            db.flush()

            updated_count += 1
            if (not was_active) and _is_record_active_today(existing, today):
                activated_ids.add(existing.id)
            continue

        record = VacationRecord(
            employee_full_name=payload.employee_full_name,
            is_special_employee=payload.is_special_employee,
            employee_positions=payload.employee_positions,
            status=payload.status,
            service=payload.service,
            deputies=imported_deputies,
            memo=payload.memo,
            start_date=payload.start_date,
            end_date=payload.end_date,
        )
        _apply_status_by_dates(record)
        db.add(record)
        db.flush()

        created_ids.append(record.id)
        all_changed_ids.add(record.id)
        if _is_record_active_today(record, today):
            activated_ids.add(record.id)

    if imported_keys_by_employee:
        open_records = db.scalars(
            select(VacationRecord).where(
                VacationRecord.status.in_(ACTIVE_STATUS_DB_VALUES),
                VacationRecord.end_date >= today,
            )
        ).all()

        for record in open_records:
            employee_key = _normalize_name_key(record.employee_full_name)
            expected_keys = imported_keys_by_employee.get(employee_key)
            if not expected_keys:
                continue
            expected_services = {item[3] for item in expected_keys}
            record_service_key = _normalize_text(record.service).lower()
            if record_service_key not in expected_services:
                continue

            record_key = _vacation_identity_key(
                record.employee_full_name,
                record.start_date,
                record.end_date,
                service=record.service,
            )
            if record_key in expected_keys:
                continue

            if record.start_date > today:
                all_changed_ids.add(record.id)
                db.delete(record)
                reconciled_deleted_count += 1
                continue

            all_changed_ids.add(record.id)
            record.status = VacationStatus.FINISHED
            record.deactivated_at = now_utc
            reconciled_archived_count += 1

    has_changes = bool(created_ids or updated_count or reconciled_archived_count or reconciled_deleted_count)
    if has_changes:
        db.commit()
        changed_ids = sorted(activated_ids)
        if changed_ids:
            event, special_names = _create_registry_notification_for_changes(db, changed_ids)
            await _broadcast_registry_update(
                changed_ids,
                event,
                special_employee_names=special_names,
            )
    else:
        db.rollback()

    max_error_items = 100
    error_items = errors[:max_error_items]
    if len(errors) > max_error_items:
        error_items.append(f"Показаны первые {max_error_items} ошибок из {len(errors)}.")

    result = {
        "created_count": len(created_ids),
        "updated_count": updated_count,
        "reconciled_archived_count": reconciled_archived_count,
        "reconciled_deleted_count": reconciled_deleted_count,
        "duplicate_count": duplicate_count,
        "skipped_empty_rows": skipped_empty_rows,
        "error_count": len(errors),
        "errors": error_items,
    }

    _add_sync_log(
        db,
        action_type="excel_import",
        source_name=filename,
        actor_username=current_user.username,
        payload={
            "created_count": len(created_ids),
            "updated_count": updated_count,
            "reconciled_archived_count": reconciled_archived_count,
            "reconciled_deleted_count": reconciled_deleted_count,
            "duplicate_count": duplicate_count,
            "skipped_empty_rows": skipped_empty_rows,
            "error_count": len(errors),
            "activated_count": len(activated_ids),
            "activated_ids": sorted(activated_ids),
            "changed_ids": sorted(all_changed_ids),
        },
    )
    db.commit()
    return result


def _build_current_vacations_statement(today: date):
    return (
        select(VacationRecord)
        .where(
            VacationRecord.status.in_(ACTIVE_STATUS_DB_VALUES),
            VacationRecord.start_date <= today,
            VacationRecord.end_date >= today,
        )
        .order_by(VacationRecord.start_date.asc(), VacationRecord.employee_full_name.asc())
    )


def _build_finished_vacations_statement(today: date):
    return (
        select(VacationRecord)
        .where(
            VacationRecord.start_date <= today,
            or_(
                VacationRecord.status == VacationStatus.FINISHED,
                VacationRecord.end_date < today,
            ),
        )
        .order_by(VacationRecord.end_date.desc(), VacationRecord.employee_full_name.asc())
    )


def _resolve_public_tab(tab: str) -> str:
    normalized = _normalize_text(tab).lower()
    if normalized not in {"current", "finished"}:
        raise HTTPException(status_code=422, detail="Параметр tab должен быть 'current' или 'finished'.")
    return normalized


@app.get("/api/vacations", response_model=list[VacationResponse])
def list_vacations(
    active_only: bool = Query(default=False, description="Только активные отпуска"),
    user: UserAccount = Depends(require_user),
    db: Session = Depends(get_db),
):
    today = datetime.now(APP_TZ).date()
    stmt = select(VacationRecord)

    if active_only:
        stmt = stmt.where(
            VacationRecord.status.in_(ACTIVE_STATUS_DB_VALUES),
            VacationRecord.start_date <= today,
            VacationRecord.end_date >= today,
        )
    else:
        # Clients see current and finished records, but not future plans.
        stmt = stmt.where(VacationRecord.start_date <= today)
        if _is_client(user):
            stmt = stmt.where(
                or_(
                    VacationRecord.status == VacationStatus.FINISHED,
                    and_(
                        VacationRecord.status.in_(ACTIVE_STATUS_DB_VALUES),
                        VacationRecord.end_date >= today,
                    ),
                )
            )

    stmt = stmt.order_by(VacationRecord.start_date.asc(), VacationRecord.employee_full_name.asc())
    return db.scalars(stmt).all()


@app.delete("/api/vacations/{record_id}", status_code=204)
async def archive_vacation(
    record_id: int,
    current_user: UserAccount = Depends(require_registry_editor),
    db: Session = Depends(get_db),
):
    record = db.get(VacationRecord, record_id)
    if not record:
        raise HTTPException(status_code=404, detail="Запись не найдена")

    record.status = VacationStatus.FINISHED
    record.deactivated_at = datetime.now(timezone.utc)
    _add_sync_log(
        db,
        action_type="manual_archive",
        actor_username=current_user.username,
        payload={"record_id": record.id, "employee_full_name": record.employee_full_name},
    )
    db.commit()
    return None


@app.delete("/api/vacations/{record_id}/hard", status_code=204)
async def hard_delete_vacation(
    record_id: int,
    current_user: UserAccount = Depends(require_admin),
    db: Session = Depends(get_db),
):
    record = db.get(VacationRecord, record_id)
    if not record:
        raise HTTPException(status_code=404, detail="Запись не найдена")

    employee_name = _normalize_text(record.employee_full_name) or f"ID {record.id}"
    db.delete(record)
    db.commit()
    _add_sync_log(
        db,
        action_type="manual_hard_delete",
        actor_username=current_user.username,
        payload={"record_id": record_id, "employee_full_name": employee_name},
    )
    db.commit()
    return None


@app.get("/api/vacations/active/export")
def export_active_vacations(_: UserAccount = Depends(require_user), db: Session = Depends(get_db)):
    today = datetime.now(APP_TZ).date()

    stmt = (
        select(VacationRecord)
        .where(
            VacationRecord.start_date <= today,
            or_(
                VacationRecord.status == VacationStatus.FINISHED,
                and_(
                    VacationRecord.status.in_(ACTIVE_STATUS_DB_VALUES),
                    VacationRecord.end_date >= today,
                ),
            ),
        )
        .order_by(VacationRecord.employee_full_name.asc())
    )
    active = db.scalars(stmt).all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Текущие и завершенные"

    headers = [
        "ФИО сотрудника",
        "Особый сотрудник",
        "Должность/Должности",
        "Статус",
        "Услуга",
        "Заместитель/и (ФИО)",
        "Фактическая должность сотрудника (замещающего)",
        "Дата начала отпуска",
        "Дата окончания отпуска",
        "Памятка",
    ]
    sheet.append(headers)

    for row in active:
        deputies_names = "\n".join(dep["deputy_full_name"] for dep in row.deputies)
        deputy_positions = "\n".join(dep["deputy_actual_position"] for dep in row.deputies)

        sheet.append(
            [
                row.employee_full_name,
                "Да" if row.is_special_employee else "Нет",
                "\n".join(row.employee_positions),
                row.status.value,
                row.service,
                deputies_names,
                deputy_positions,
                row.start_date.isoformat(),
                row.end_date.isoformat(),
                row.memo or "",
            ]
        )

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    filename = f"vacations_current_and_finished_{today.isoformat()}.xlsx"
    headers = {"Content-Disposition": f"attachment; filename={filename}"}
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.get("/api/public/vacations", response_model=list[VacationResponse])
def public_vacations(
    tab: str = Query(default="current", description="current | finished"),
    db: Session = Depends(get_db),
):
    today = datetime.now(APP_TZ).date()
    resolved_tab = _resolve_public_tab(tab)
    stmt = _build_current_vacations_statement(today) if resolved_tab == "current" else _build_finished_vacations_statement(today)
    return db.scalars(stmt).all()


@app.get("/api/public/vacations/export")
def public_export_vacations(
    tab: str = Query(default="current", description="current | finished"),
    db: Session = Depends(get_db),
):
    today = datetime.now(APP_TZ).date()
    resolved_tab = _resolve_public_tab(tab)
    stmt = _build_current_vacations_statement(today) if resolved_tab == "current" else _build_finished_vacations_statement(today)
    rows = db.scalars(stmt).all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Текущие отпуска" if resolved_tab == "current" else "Завершенные случаи"
    sheet.append(
        [
            "ФИО сотрудника",
            "Особый сотрудник",
            "Должность/Должности",
            "Статус",
            "Услуга",
            "Заместитель/и (ФИО)",
            "Фактическая должность сотрудника (замещающего)",
            "Дата начала отпуска",
            "Дата окончания отпуска",
            "Памятка",
        ]
    )

    for row in rows:
        sheet.append(
            [
                row.employee_full_name,
                "Да" if row.is_special_employee else "Нет",
                "\n".join(row.employee_positions),
                row.status.value,
                row.service,
                "\n".join(dep.get("deputy_full_name", "") for dep in row.deputies or []),
                "\n".join(dep.get("deputy_actual_position", "") for dep in row.deputies or []),
                row.start_date.isoformat(),
                row.end_date.isoformat(),
                row.memo or "",
            ]
        )

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    suffix = "current" if resolved_tab == "current" else "finished"
    filename = f"vacations_{suffix}_{today.isoformat()}.xlsx"
    headers = {"Content-Disposition": f"attachment; filename={filename}"}
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.websocket("/ws/registry")
async def registry_ws(websocket: WebSocket):
    await hub.connect(websocket)
    try:
        while True:
            await websocket.receive_text()
    except WebSocketDisconnect:
        hub.disconnect(websocket)
    except Exception:
        hub.disconnect(websocket)

